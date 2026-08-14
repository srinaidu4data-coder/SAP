"""1-minute Analyze: Excel + rules + report from proven live counts.

Live GUI cannot open 300 tables in 60 seconds. This path uses every F7
already proven on this book, the process catalog, and field values read
on the glass (T001, TVAK). Parallel agents write rules, integrations,
and the consultant report while the optional glass walk continues.
"""
from __future__ import annotations

from concurrent.futures import ThreadPoolExecutor
from typing import Any

from sapilot.product.drillthrough import layer_of
from sapilot.product.table_read import _MEANING

# Field / config values actually seen on this book (glass), not catalog guesses.
_GLASS_FIELDS: list[dict[str, str]] = [
    {"layer": "config", "table": "TVAK", "field": "AUART", "value": "01, 5OR, 5ORT, 5RE, AA, AD1, AD2, AD3, AD9, AE, AEBO", "note": "Sales document types on the open TVAK list"},
    {"layer": "config", "table": "T001", "field": "BUKRS", "value": "0001 BRIDGE ANIS, 0002 Porite India, 0003 SAP US, 0004 A2Z, 001 Adrita, 0070 Company FI, 0079 Vijay CC", "note": "Company codes on the T001 list"},
    {"layer": "config", "table": "TCK03", "field": "KLVAR", "value": "74 costing variants", "note": "F7 on this book"},
    {"layer": "config", "table": "TCK31", "field": "overhead sheet", "value": "1", "note": "Almost unused overhead"},
    {"layer": "master", "table": "KNKK", "field": "credit master", "value": "0 rows", "note": "No classic credit master. 12k orders still exist."},
    {"layer": "master", "table": "KNA1", "field": "KUNNR", "value": "1,284 customers (Execute hits)", "note": "Customer general"},
    {"layer": "master", "table": "KNVV", "field": "sales area", "value": "795", "note": "F7"},
    {"layer": "transaction", "table": "NAST", "field": "output", "value": "928 messages", "note": "Most invoices never produced a letter"},
    {"layer": "transaction", "table": "EDIDC", "field": "IDoc", "value": "1,239", "note": "Interface clock is live"},
    {"layer": "transaction", "table": "FARR_D_CONTRACT", "field": "RAR contract", "value": "0 / empty", "note": "Table exists and is empty"},
    {"layer": "transaction", "table": "FARR_D_REVENUE", "field": "RAR revenue", "value": "does not exist", "note": "SAP: check the name"},
]


def _n(counts: dict, tab: str) -> int | None:
    rec = counts.get((tab or "").upper())
    if not rec:
        return None
    return rec.get("entries_found")


def rules_agent(counts: dict, asked: str) -> list[dict[str, str]]:
    """Business rules this book is actually running."""
    vbak, likp, vbrk = _n(counts, "VBAK"), _n(counts, "LIKP"), _n(counts, "VBRK")
    bsid, bsad, nast = _n(counts, "BSID"), _n(counts, "BSAD"), _n(counts, "NAST")
    knkk, farr = _n(counts, "KNKK"), _n(counts, "FARR_D_CONTRACT")
    out: list[dict[str, str]] = []
    if vbak is not None and likp is not None:
        out.append({
            "rule": "Delivery is optional after order",
            "evidence": f"VBAK={vbak:,} LIKP={likp:,}. {vbak - likp:,} orders never shipped.",
            "requirement": "The commercial process allows order without goods issue. That is a design, not a bug.",
        })
    if likp is not None and vbrk is not None:
        out.append({
            "rule": "Billing can lag delivery",
            "evidence": f"LIKP={likp:,} VBRK={vbrk:,}. {likp - vbrk:,} deliveries unbilled.",
            "requirement": "Unbilled delivery is allowed. RAR cannot start there.",
        })
    if knkk == 0 and vbak:
        out.append({
            "rule": "Classic credit is not a gate",
            "evidence": f"KNKK=0 while VBAK={vbak:,}.",
            "requirement": "No KNKK limit is required to sell. Credit is a decision, not a missing t-code.",
        })
    if vbrk and nast is not None:
        out.append({
            "rule": "Invoice output is not mandatory",
            "evidence": f"VBRK={vbrk:,} NAST={nast:,}.",
            "requirement": "Most bills do not create a customer message. Collections call people about invoices they never saw.",
        })
    if farr == 0 and vbrk:
        out.append({
            "rule": "IFRS 15 / RAR is not the recognition clock",
            "evidence": "FARR_D_CONTRACT empty. FARR_D_REVENUE is not a table here. VBRK is full.",
            "requirement": "The process that is possible is classic SD billing → FI. Not POB fulfillment.",
        })
    if bsid is not None and bsad is not None:
        out.append({
            "rule": "Cash collection is a second clock",
            "evidence": f"BSID open={bsid:,} BSAD cleared={bsad:,}.",
            "requirement": "Clearing happens when someone works it. That is not revenue recognition.",
        })
    if _n(counts, "TCK31") == 1 and _n(counts, "TCK03"):
        out.append({
            "rule": "Costing recipes outnumber used overhead",
            "evidence": f"TCK03={_n(counts, 'TCK03')} variants, TCK31=1 sheet.",
            "requirement": "Margin on the invoice is only as honest as the one overhead sheet actually used.",
        })
    if not out:
        out.append({
            "rule": "No live census yet",
            "evidence": asked,
            "requirement": "Log into SAP and prove Number of Entries. Rules are written from glass facts.",
        })
    return out


def integrations_agent(counts: dict) -> list[dict[str, str]]:
    out = []
    if _n(counts, "NAST") is not None:
        out.append({"interface": "NAST output", "n": _n(counts, "NAST"), "meaning": "Print/EDI/email messages that actually exist."})
    if _n(counts, "EDIDC") is not None:
        out.append({"interface": "IDoc EDIDC", "n": _n(counts, "EDIDC"), "meaning": "Inbound/outbound IDocs on this client."})
    if _n(counts, "EDIDS") is not None:
        out.append({"interface": "IDoc status EDIDS", "n": _n(counts, "EDIDS"), "meaning": "IDoc status records."})
    if _n(counts, "FARR_D_CONTRACT") == 0:
        out.append({"interface": "RAR / RAI", "n": 0, "meaning": "No RAR contracts. Billing is not posting into FARR."})
    if _n(counts, "VBFA"):
        out.append({"interface": "Document flow VBFA", "n": _n(counts, "VBFA"), "meaning": "Document hops stored in the book."})
    return out


def process_possible(counts: dict, asked: str, rec: dict) -> list[str]:
    lines = [
        rec.get("spine") or "",
        rec.get("story") or "",
    ]
    vbrk, farr = _n(counts, "VBRK"), _n(counts, "FARR_D_CONTRACT")
    if vbrk and farr == 0:
        lines.append(
            "Possible on this book: order → delivery → billing → AR (BSID/BSAD). "
            "Not possible: RAR contract → POB → fulfillment → IFRS 15 revenue table."
        )
    if _n(counts, "KNKK") == 0:
        lines.append("Possible: sell with no KNKK credit master. Not possible to claim credit-check was the gate.")
    if _n(counts, "KEKO"):
        lines.append("Possible: costed sales (KEKO live). Invoice margin can be challenged against estimates.")
    return [x for x in lines if x]


def write_minute_excel(path: str, *, asked: str, counts: dict, drill: dict, rules: list, integ: list, possible: list) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    fill = PatternFill("solid", fgColor="1F4E79")
    hf = Font(bold=True, color="FFFFFF")

    def head(ws, cols, row=1):
        for c, h in enumerate(cols, 1):
            cell = ws.cell(row, c, h)
            cell.font = hf
            cell.fill = fill

    idx = wb.active
    idx.title = "Index"
    idx["A1"] = f"Analyze — {asked}"
    idx["A1"].font = Font(bold=True, size=16)
    idx["A2"] = "1-minute pack. Live F7 on this book + glass field values. Not a 12k-row OCR dump."
    idx["A4"] = "Sheets: Transactional, Master, Configuration, Field values, Business rules, Integrations, Process possible, All tables"

    layered = {"transaction": [], "master": [], "config": []}
    for rec in counts.values():
        tab = rec.get("table") or ""
        layered.setdefault(layer_of(tab), []).append(
            {"table": tab, "n": rec.get("entries_found"), "rank": rec.get("rank"), "layer": layer_of(tab)}
        )

    for title, key in (("Transactional", "transaction"), ("Master", "master"), ("Configuration", "config")):
        ws = wb.create_sheet(title)
        head(ws, ["Table", "Layer", "Entries (F7)", "Meaning"])
        rows = layered.get(key) or layered.get({"Transactional": "transaction", "Master": "master", "Configuration": "config"}[title]) or []
        if not rows:
            # fall back to classifying all counts
            rows = [{"table": c["table"], "n": c.get("entries_found"), "layer": layer_of(c["table"])} for c in counts.values() if layer_of(c["table"]) == key]
        for i, r in enumerate(rows, 2):
            tab = r.get("table") or ""
            ws.cell(i, 1, tab)
            ws.cell(i, 2, r.get("layer") or key)
            ws.cell(i, 3, r.get("n"))
            ws.cell(i, 4, _MEANING.get(tab, ""))

    fv = wb.create_sheet("Field values")
    head(fv, ["Layer", "Table", "Field", "Value on this book", "Note"])
    for i, r in enumerate(_GLASS_FIELDS, 2):
        fv.cell(i, 1, r["layer"])
        fv.cell(i, 2, r["table"])
        fv.cell(i, 3, r["field"])
        fv.cell(i, 4, r["value"])
        fv.cell(i, 5, r["note"])

    br = wb.create_sheet("Business rules")
    head(br, ["Rule this book is running", "Evidence", "Business requirement"])
    for i, r in enumerate(rules, 2):
        br.cell(i, 1, r.get("rule"))
        br.cell(i, 2, r.get("evidence"))
        br.cell(i, 3, r.get("requirement"))

    ig = wb.create_sheet("Integrations")
    head(ig, ["Interface", "Count", "Meaning"])
    for i, r in enumerate(integ, 2):
        ig.cell(i, 1, r.get("interface"))
        ig.cell(i, 2, r.get("n"))
        ig.cell(i, 3, r.get("meaning"))

    pp = wb.create_sheet("Process possible")
    pp["A1"] = "What process is possible on this book"
    pp["A1"].font = Font(bold=True)
    for i, line in enumerate(possible, 3):
        pp.cell(i, 1, line)

    alls = wb.create_sheet("All tables")
    head(alls, ["Table", "Layer", "Entries", "Rank", "Notes"])
    for i, rec in enumerate(sorted(counts.values(), key=lambda x: x.get("table") or ""), 2):
        tab = rec.get("table") or ""
        alls.cell(i, 1, tab)
        alls.cell(i, 2, layer_of(tab))
        alls.cell(i, 3, rec.get("entries_found"))
        alls.cell(i, 4, rec.get("rank"))
        alls.cell(i, 5, rec.get("notes") or "")

    wb.save(path)
    return path


def run_agents(counts: dict, asked: str, rec: dict) -> dict[str, Any]:
    with ThreadPoolExecutor(max_workers=3) as pool:
        f_rules = pool.submit(rules_agent, counts, asked)
        f_int = pool.submit(integrations_agent, counts)
        f_pos = pool.submit(process_possible, counts, asked, rec)
        return {
            "rules": f_rules.result(),
            "integrations": f_int.result(),
            "possible": f_pos.result(),
        }
