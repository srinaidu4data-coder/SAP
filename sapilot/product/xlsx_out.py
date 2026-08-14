"""Write a .xlsx with no extra packages. Excel opens it."""
from __future__ import annotations

import zipfile
from pathlib import Path
from xml.sax.saxutils import escape


def _col(n: int) -> str:
    s = ""
    x = n
    while x:
        x, r = divmod(x - 1, 26)
        s = chr(65 + r) + s
    return s or "A"


def write_xlsx(path: str | Path, headers: list[str], rows: list[list[str]], *, sheet: str = "Sheet1") -> str:
    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    heads = [str(h or "") for h in (headers or [])]
    body = [[str(c or "") for c in r] for r in (rows or [])]
    width = max(len(heads), max((len(r) for r in body), default=0), 1)
    while len(heads) < width:
        heads.append(f"C{len(heads) + 1}")

    def cell(r: int, c: int, val: str, header: bool = False) -> str:
        ref = f"{_col(c + 1)}{r + 1}"
        t = escape(val)
        style = ' s="1"' if header else ""
        return f'<c r="{ref}" t="inlineStr"{style}><is><t>{t}</t></is></c>'

    lines = ['<row r="1">']
    lines.extend(cell(0, i, h, True) for i, h in enumerate(heads))
    lines.append("</row>")
    for ri, row in enumerate(body, start=1):
        lines.append(f'<row r="{ri + 1}">')
        padded = list(row) + [""] * (width - len(row))
        lines.extend(cell(ri, i, padded[i]) for i in range(width))
        lines.append("</row>")
    sheet_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f"<sheetData>{''.join(lines)}</sheetData></worksheet>"
    )
    workbook = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        f'<sheets><sheet name="{escape(sheet[:31]) or "Sheet1"}" sheetId="1" r:id="rId1"/></sheets>'
        "</workbook>"
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
        "</Relationships>"
    )
    wb_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
        "</Relationships>"
    )
    ctypes = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        "</Types>"
    )
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("[Content_Types].xml", ctypes)
        z.writestr("_rels/.rels", rels)
        z.writestr("xl/workbook.xml", workbook)
        z.writestr("xl/_rels/workbook.xml.rels", wb_rels)
        z.writestr("xl/worksheets/sheet1.xml", sheet_xml)
    return str(out)


def write_job_workbook(job: dict) -> str:
    """One workbook for a research sitting. Only rows that were on the glass."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    dest = Path(job.get("dir") or ".")
    dest.mkdir(parents=True, exist_ok=True)
    path = dest / "TABLES.xlsx"
    wb = Workbook()
    cover = wb.active
    cover.title = "Index"
    cover["A1"] = job.get("title") or job.get("asked") or "Tables"
    cover["A1"].font = Font(bold=True, size=14)
    cover["A2"] = "A tab is written only after Execute. Empty tab = not opened."
    fill = PatternFill("solid", fgColor="1F4E79")
    hf = Font(bold=True, color="FFFFFF")
    cover["A4"] = "Table"
    cover["B4"] = "F7 count"
    cover["C4"] = "Opened"
    cover["D4"] = "Glass rows"
    for cell in (cover["A4"], cover["B4"], cover["C4"], cover["D4"]):
        cell.font = hf
        cell.fill = fill
    for i, rec in enumerate(job.get("counts") or [], 5):
        table = rec.get("table") or "?"
        cover.cell(i, 1, table)
        cover.cell(i, 2, rec.get("entries_found"))
        cover.cell(i, 3, "yes" if rec.get("opened") else "no")
        contents = rec.get("contents") or {}
        rows = contents.get("sample_rows") or []
        cover.cell(i, 4, contents.get("visible_rows") or len(rows))
        ws = wb.create_sheet(str(table)[:31])
        ws["A1"] = table
        ws["B1"] = contents.get("story") or rec.get("notes") or ""
        ws["A2"] = f"F7={rec.get('entries_found')} opened={bool(rec.get('opened'))}"
        cols = contents.get("columns") or []
        for c, h in enumerate(cols, 1):
            cell = ws.cell(4, c, h)
            cell.font = hf
            cell.fill = fill
        for r, row in enumerate(rows, 5):
            for c, val in enumerate(row, 1):
                ws.cell(r, c, val)
        if not rec.get("opened"):
            ws["A4"] = "Not executed — no rows."
    wb.save(path)
    return str(path)
